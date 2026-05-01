import os, tempfile
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from app.database import get_db
from app.models.patient import Patient
from app.models.lab_test import LabTest
from app.models.tcm_score import TcmScore
from app.models.quality_of_life import QualityOfLife
from app.models.treatment import Treatment
from app.models.alert import Alert
from app.services.pdf_service import generate_patient_pdf
from app.models.treatment import Treatment
from app.models.alert import Alert
from app.services.pdf_service import generate_patient_pdf

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/overview")
def overview(db: Session = Depends(get_db)):
    total_patients = db.query(Patient).count()
    pending_alerts = db.query(Alert).filter(Alert.status == "pending").count()
    high_alerts = db.query(Alert).filter(Alert.level == "high", Alert.status == "pending").count()
    return {
        "total_patients": total_patients,
        "pending_alerts": pending_alerts,
        "high_alerts": high_alerts,
    }


@router.get("/trend/{patient_id}")
def patient_trend(patient_id: int, db: Session = Depends(get_db)):
    tcm_scores = db.query(TcmScore).filter(TcmScore.patient_id == patient_id).order_by(TcmScore.record_date).all()
    lab_tests = db.query(LabTest).filter(LabTest.patient_id == patient_id).order_by(LabTest.record_date).all()
    qol_records = db.query(QualityOfLife).filter(QualityOfLife.patient_id == patient_id).order_by(QualityOfLife.record_date).all()
    return {
        "tcm_scores": [{"date": str(s.record_date), "total_score": s.total_score} for s in tcm_scores],
        "lab_tests": [{"date": str(t.record_date), "type": t.test_type, "alt": t.alt, "hgb": t.hgb, "wbc": t.wbc} for t in lab_tests],
        "qol": [{"date": str(q.record_date), "total_score": q.total_score} for q in qol_records],
    }


# === PDF 导出 ===
@router.get("/patient/{id}/pdf")
def export_patient_pdf(id: int, db: Session = Depends(get_db)):
    patient = db.query(Patient).filter(Patient.id == id).first()
    if not patient:
        raise HTTPException(404, "患者不存在")
    tcm = db.query(TcmScore).filter(TcmScore.patient_id == id).order_by(TcmScore.record_date).all()
    lab = db.query(LabTest).filter(LabTest.patient_id == id).order_by(LabTest.record_date).all()
    qol = db.query(QualityOfLife).filter(QualityOfLife.patient_id == id).order_by(QualityOfLife.record_date).all()
    tx = db.query(Treatment).filter(Treatment.patient_id == id).order_by(Treatment.start_date).all()
    alerts = db.query(Alert).filter(Alert.patient_id == id).order_by(Alert.created_at.desc()).all()
    tmp = tempfile.mktemp(suffix=".pdf")
    generate_patient_pdf(patient, tcm, lab, qol, tx, alerts, tmp)
    return FileResponse(tmp, media_type="application/pdf", filename=f"患者报告_{patient.name}.pdf")


# === Excel 导出 ===
@router.get("/export/patients")
def export_patients_excel(db: Session = Depends(get_db)):
    patients = db.query(Patient).order_by(Patient.admission_date.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "患者列表"
    headers = ["ID", "姓名", "性别", "年龄", "电话", "入院日期", "西医诊断", "中医诊断", "过敏史", "既往史", "家族史", "入院评估", "出院小结", "备注"]
    ws.append(headers)
    for p in patients:
        ws.append([p.id, p.name, p.gender, p.age, p.phone, str(p.admission_date),
                    p.diagnosis, p.tcm_diagnosis, p.allergy_history, p.past_history,
                    p.family_history, p.admission_assessment, p.discharge_summary, p.notes])
    tmp = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp)
    return FileResponse(tmp, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="患者列表.xlsx")


@router.get("/export/lab-tests/{patient_id}")
def export_lab_excel(patient_id: int, db: Session = Depends(get_db)):
    labs = db.query(LabTest).filter(LabTest.patient_id == patient_id).order_by(LabTest.record_date).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "检验数据"
    headers = ["日期", "类型", "WBC", "RBC", "HGB", "PLT", "ALT", "AST", "TBIL", "ALB", "胃泌素", "PGI", "PGII"]
    ws.append(headers)
    for t in labs:
        ws.append([str(t.record_date), t.test_type, t.wbc, t.rbc, t.hgb, t.plt,
                    t.alt, t.ast, t.tbil, t.alb, t.gastrin, t.pepsinogen_i, t.pepsinogen_ii])
    tmp = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp)
    return FileResponse(tmp, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="检验数据.xlsx")


# === Excel 导入 ===
from fastapi import UploadFile, File
from openpyxl import load_workbook

@router.post("/import/patients")
def import_patients_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    wb = load_workbook(filename=file.file)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        p = Patient(
            name=row[1], gender=row[2] or "男", age=row[3] or 0, phone=row[4] or "",
            admission_date=row[5] if row[5] else "2024-01-01",
            diagnosis=row[6] or "", tcm_diagnosis=row[7] or "",
            allergy_history=row[8] or "", past_history=row[9] or "",
            family_history=row[10] or "", admission_assessment=row[11] or "",
            discharge_summary=row[12] or "", notes=row[13] or "",
        )
        db.add(p)
        count += 1
    db.commit()
    return {"message": f"成功导入 {count} 条患者数据"}


# === 疗效评价 ===
@router.get("/efficacy/{patient_id}")
def efficacy_evaluation(patient_id: int, db: Session = Depends(get_db)):
    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(404, "患者不存在")
    tcm = db.query(TcmScore).filter(TcmScore.patient_id == patient_id).order_by(TcmScore.record_date).all()
    lab = db.query(LabTest).filter(LabTest.patient_id == patient_id).order_by(LabTest.record_date).all()
    qol = db.query(QualityOfLife).filter(QualityOfLife.patient_id == patient_id).order_by(QualityOfLife.record_date).all()

    result = {"patient_id": patient_id, "patient_name": patient.name}

    # 中医证候疗效：积分变化率
    if len(tcm) >= 2:
        first_score = tcm[0].total_score
        last_score = tcm[-1].total_score
        change_rate = ((first_score - last_score) / first_score * 100) if first_score > 0 else 0
        result["tcm_efficacy"] = {
            "first_score": first_score, "last_score": last_score,
            "change_rate": round(change_rate, 2),
            "improvement": change_rate >= 30,
            "scores_timeline": [{"date": str(s.record_date), "total": s.total_score} for s in tcm],
        }

    # 实验室指标疗效
    if len(lab) >= 2:
        first_lab = lab[0]
        last_lab = lab[-1]
        lab_changes = {}
        for field in ["alt", "ast", "hgb", "wbc", "tbil", "alb"]:
            fv = getattr(first_lab, field, None)
            lv = getattr(last_lab, field, None)
            if fv and lv:
                change = ((fv - lv) / fv * 100) if fv != 0 else 0
                lab_changes[field] = {"first": fv, "last": lv, "change_rate": round(change, 2)}
        result["lab_efficacy"] = lab_changes

    # 生活质量疗效
    if len(qol) >= 2:
        first_qol = qol[0].total_score
        last_qol = qol[-1].total_score
        qol_change = ((last_qol - first_qol) / first_qol * 100) if first_qol > 0 else 0
        result["qol_efficacy"] = {
            "first_score": first_qol, "last_score": last_qol,
            "change_rate": round(qol_change, 2),
        }

    return result